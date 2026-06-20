# pipeline/transfer.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# 은행 코드표 (대부분의 은행 양식에서 공통 사용)
BANK_CODES = {
    "국민": "004", "신한": "088", "우리": "020", "하나": "081",
    "농협": "011", "기업": "003", "카카오": "090", "토스": "092",
    "대구": "031", "부산": "032", "경남": "039", "광주": "034",
    "전북": "037", "제주": "035", "새마을": "045", "신협": "048",
    "우체국": "071", "SC": "023", "씨티": "027",
}

def build_transfer_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    분석 완료된 df에서 이체 대상자 행만 추출하여
    대량이체 양식에 맞는 DataFrame 반환
    """
    # 수당이 0원인 미지급 대상자 제외
    target = df[df["total_allow"] > 0].copy()
    target = target.sort_values("participant_name").reset_index(drop=True)

    transfer_df = pd.DataFrame({
        "순번":             range(1, len(target) + 1),
        "받는분_예금주":    target["participant_name"],
        "받는분_계좌번호":  "",           # 계좌 정보 없음 → 빈칸
        "받는분_은행":      "",           # 계좌 정보 없음 → 빈칸
        "이체금액":         target["total_allow"],
        "받는분_통장메모":  "미래내일일경험수당",
        "보내는분_통장메모": target["participant_name"] + " 수당",
        # 참고용 컬럼 (실제 이체 양식에는 포함 안 해도 됨)
        "_직무훈련수당":    target["jt_allow"],
        "_일경험수당":      target["we_allow"],
        "_출석률":          target["avg_rate"].round(1),
        "_프로그램":        target["program_name"],
    })

    return transfer_df


def export_transfer_excel(
    df: pd.DataFrame,
    output_path: str,
    template_path: str = None
) -> str:
    """
    대량이체 엑셀 파일 생성
    - template_path 있으면 양식 파일에 데이터 삽입
    - 없으면 기본 스타일로 새로 생성
    """
    transfer_df = build_transfer_rows(df)
    today = datetime.today().strftime("%Y%m%d")
    filename = f"{output_path}/대량이체_{today}.xlsx"

    if template_path:
        _write_to_template(transfer_df, template_path, filename)
    else:
        _write_new(transfer_df, filename)

    total_amount = transfer_df["이체금액"].sum()
    total_count  = len(transfer_df)
    excluded     = len(df) - total_count

    print(f"✅ 대량이체 파일 생성: {filename}")
    print(f"   이체 대상: {total_count}명 / 제외(미지급): {excluded}명")
    print(f"   총 이체금액: {total_amount:,}원")

    return filename


def _write_to_template(transfer_df, template_path, output_path):
    """은행 제공 양식에 데이터 삽입 (양식 서식 유지)"""
    wb = load_workbook(template_path)
    ws = wb.active

    # 대부분 양식의 데이터 시작 행 = 4행 (헤더 3줄 후)
    # 실제 양식 확인 후 START_ROW 수정 필요
    START_ROW = 4
    COLS = ["받는분_예금주", "받는분_계좌번호", "받는분_은행",
            "이체금액", "받는분_통장메모", "보내는분_통장메모"]

    for i, row in transfer_df.iterrows():
        for j, col in enumerate(COLS):
            ws.cell(row=START_ROW + i, column=j + 1, value=row[col])

    wb.save(output_path)


def _write_new(transfer_df, output_path):
    """양식 없을 때 기본 스타일로 새로 생성"""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ① 이체 시트 (실제 업로드용)
        upload_cols = [
            "순번", "받는분_예금주", "받는분_계좌번호",
            "받는분_은행", "이체금액", "받는분_통장메모", "보내는분_통장메모"
        ]
        transfer_df[upload_cols].to_excel(
            writer, sheet_name="대량이체", index=False
        )

        # ② 상세 내역 시트 (운영자 확인용)
        detail_cols = [
            "순번", "받는분_예금주", "_프로그램", "_출석률",
            "_직무훈련수당", "_일경험수당", "이체금액"
        ]
        transfer_df[detail_cols].rename(columns={
            "_프로그램": "프로그램", "_출석률": "출석률(%)",
            "_직무훈련수당": "직무훈련수당", "_일경험수당": "일경험수당"
        }).to_excel(writer, sheet_name="수당상세내역", index=False)

        # ③ 요약 시트
        summary = pd.DataFrame({
            "항목": ["이체 대상 인원", "총 이체금액", "직무훈련 수당 합계",
                     "일경험 수당 합계", "생성 일시"],
            "값": [
                f"{len(transfer_df)}명",
                f"{transfer_df['이체금액'].sum():,}원",
                f"{transfer_df['_직무훈련수당'].sum():,}원",
                f"{transfer_df['_일경험수당'].sum():,}원",
                datetime.today().strftime("%Y-%m-%d %H:%M"),
            ]
        })
        summary.to_excel(writer, sheet_name="요약", index=False)

    _apply_style(output_path)


def _apply_style(path):
    """헤더 색상·정렬·열너비 적용"""
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1A2035")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        # 헤더 스타일
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 열 너비 자동 조정
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value else 0) for c in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # 이체금액 열 숫자 포맷
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 1000:
                    cell.number_format = "#,##0"
                cell.border = border

    wb.save(path)